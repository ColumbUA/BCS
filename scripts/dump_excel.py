import openpyxl, json
wb = openpyxl.load_workbook('/app/output/BCHS.xlsx', data_only=True)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"=== SHEET: {s}  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column} ===")
    for r in ws.iter_rows(values_only=False):
        for c in r:
            if c.value not in (None, ""):
                print(f"  {c.coordinate}: {repr(c.value)}")
